#!/usr/bin/env python3
# version: 6
"""
Comeet → Greenhouse candidate import transformer.

Usage:
    python3 transform.py <comeet_export_file> [--out <output_file>]

Accepts:  .numbers (Apple Numbers), .csv, .xlsx / .xlsm
Outputs:  greenhouse_[original name].xlsx  (same folder as input unless --out given)
"""

import argparse
import re
import sys
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependencies. Run: pip3 install pandas openpyxl numbers-parser")

GREENHOUSE_COLUMNS = [
    "First Name", "Last Name", "Company", "Title", "Notes",
    "Email", "Phone", "Social Media", "Website", "Country",
    "Source", "Who gets credit", "Job", "Milestone",
    "Salary Expectations",
    "Last Completed Step (Recruit)",
    "Current Stage (Recruit)",
    "Candidate Current Status (Recruit)",
    "Disposition Date (Recruit)",
    "Disposition Reason (Recruit)",
    "Disposition Notes (Recruit)",
    "Original Application Date (Recruit)",
]

# ── Milestone mapping ─────────────────────────────────────────────────────────
_MILESTONE_RULES = [
    ("Hired",        ["hired", "onboard"]),
    ("Offer",        ["offer"]),
    ("Face to Face", ["face to face", "in person", "in-person", "onsite", "on-site"]),
    ("Assessment",   ["phone", "video", "introduct", "interview", "technical",
                      "assignment", "exercise", "test", "task", "home assign", "take home"]),
]

def normalize_milestone(stage: str) -> str:
    if not stage:
        return "Application"
    lower = stage.lower()
    for milestone, keywords in _MILESTONE_RULES:
        if any(kw in lower for kw in keywords):
            return milestone
    return "Application"


# ── helpers ───────────────────────────────────────────────────────────────────

def _val(row, col):
    """Return stripped string or '' for None/NaN. Handles duplicate column names."""
    import math

    v = row.get(col) if isinstance(row, dict) else (
        row[col] if col in row.index else None
    )
    if isinstance(v, pd.Series):
        non_null = v.dropna()
        v = non_null.iloc[0] if len(non_null) > 0 else None

    if v is None:
        return ""
    try:
        if isinstance(v, float) and math.isnan(v):
            return ""
    except Exception:
        pass
    s = str(v).strip()
    return "" if s.lower() in ("none", "nan") else s


def clean_phone(val) -> str:
    if val is None:
        return ""
    import math
    try:
        if isinstance(val, float) and math.isnan(val):
            return ""
    except Exception:
        pass
    s = str(val).strip()
    if s.lower() in ("none", "nan"):
        return ""
    if s.endswith(".0"):
        s = s[:-2]
    # Strip leading apostrophe (some Comeet exports prefix phones with ')
    s = s.lstrip("'")
    return s


def split_name(name: str):
    if not name:
        return "", ""
    parts = name.strip().split(" ", 1)
    return parts[0], (parts[1] if len(parts) > 1 else "")


def _clean_title(col: str) -> str:
    """Clean a column header for use as a Notes label: remove bullet/tab chars."""
    cleaned = re.sub(r'[•\t]+', ' ', col).strip()
    return re.sub(r'\s+', ' ', cleaned)


def build_notes(row, all_columns: list) -> str:
    """
    Notes structure:
      1. One line per non-empty field from column AE ("Current step/s") onwards,
         format:  Title: Value
         (Skills column is skipped here and appended at the end instead)
      2. Skills: <value>  (always last)
    """
    # Find the starting column dynamically; fall back to index 30 (Excel col AE)
    try:
        ae_start = all_columns.index("Current step participants")
    except ValueError:
        ae_start = 32

    lines = []
    seen: set = set()

    for col in all_columns[ae_start:]:
        if not col or col in seen:
            continue
        seen.add(col)

        if col == "Skills":        # handled at the end
            continue

        v = _val(row, col)
        if not v:
            continue

        lines.append(f"{_clean_title(col)}: {v}")

    # Skills at the bottom
    skills = _val(row, "Skills")
    if skills:
        lines.append(f"Skills: {skills}")

    return "\n".join(lines)


# ── file readers ──────────────────────────────────────────────────────────────

def read_numbers(path: Path) -> pd.DataFrame:
    try:
        from numbers_parser import Document
    except ImportError:
        sys.exit("numbers-parser not installed. Run: pip3 install numbers-parser")
    doc = Document(str(path))
    best_table, best_row_count = None, -1
    for sheet in doc.sheets:
        for table in sheet.tables:
            if table.num_rows > best_row_count:
                best_row_count = table.num_rows
                best_table = table
    if best_table is None:
        sys.exit("No tables found in the .numbers file.")
    rows    = list(best_table.iter_rows())
    headers = [c.value for c in rows[0]]
    data    = [[c.value for c in row] for row in rows[1:]]
    return pd.DataFrame(data, columns=headers)


def read_input(path: Path) -> pd.DataFrame:
    ext = path.suffix.lower()
    if ext == ".numbers":
        return read_numbers(path)
    elif ext == ".csv":
        return pd.read_csv(path, dtype=str)
    elif ext in (".xlsx", ".xlsm", ".xls"):
        return pd.read_excel(path, dtype=str)
    else:
        sys.exit(f"Unsupported file type: {ext!r}. Expected .numbers, .csv, or .xlsx")


# ── transform ─────────────────────────────────────────────────────────────────

def transform(df: pd.DataFrame) -> pd.DataFrame:
    all_cols = list(df.columns)
    output_rows = []

    for _, row in df.iterrows():
        first, last = split_name(_val(row, "Name"))

        output_rows.append({
            "First Name":                        first,
            "Last Name":                         last,
            "Company":                           _val(row, "Current company"),
            "Title":                             _val(row, "Current position"),
            "Notes":                             build_notes(row, all_cols),
            "Email":                             _val(row, "Email"),
            "Phone":                             clean_phone(
                                                     row.get("Phone #1") if isinstance(row, dict)
                                                     else (row["Phone #1"] if "Phone #1" in row.index else None)
                                                 ),
            "Social Media":                      _val(row, "Candidate LinkedIn URL"),
            "Website":                           "",
            "Country":                           _val(row, "Country"),
            "Source":                            _val(row, "Source Name"),
            "Who gets credit":                   _val(row, "Recruiter(s)").split(",")[0].strip(),
            "Job":                               _val(row, "Position Name"),
            "Milestone":                         normalize_milestone(_val(row, "Current stage")),
            "Salary Expectations":               _val(row, "Salary expectations"),
            "Last Completed Step (Recruit)":     _val(row, "Last Completed Step"),
            "Current Stage (Recruit)":           _val(row, "Current stage"),
            "Candidate Current Status (Recruit)": _val(row, "Candidate Current Status"),
            "Disposition Date (Recruit)":        _val(row, "Last Candidate Status Change Date (yyyy-MM-dd HH:mm)"),
            "Disposition Reason (Recruit)":      _val(row, "Disposition Reason"),
            "Disposition Notes (Recruit)":       _val(row, "Disposition Notes"),
            "Original Application Date (Recruit)": _val(row, "Applied Date (yyyy-MM-dd HH:mm)"),
        })

    return pd.DataFrame(output_rows, columns=GREENHOUSE_COLUMNS)


# ── Excel formatting ──────────────────────────────────────────────────────────

COL_WIDTHS = {
    "First Name": 18,  "Last Name": 20,   "Company": 28,  "Title": 30,
    "Notes": 65,       "Email": 30,        "Phone": 18,   "Social Media": 35,
    "Website": 20,     "Country": 18,      "Source": 18,  "Who gets credit": 22,
    "Job": 30,         "Milestone": 18,    "Salary Expectations": 22,
    "Last Completed Step (Recruit)": 28,
    "Current Stage (Recruit)": 24,
    "Candidate Current Status (Recruit)": 28,
    "Disposition Date (Recruit)": 24,
    "Disposition Reason (Recruit)": 28,
    "Disposition Notes (Recruit)": 40,
    "Original Application Date (Recruit)": 28,
}

HEADER_BG   = "1F4E79"
HEADER_FONT = "FFFFFF"


def format_workbook(path: Path):
    wb = load_workbook(str(path))
    ws = wb["Sheet people"]

    for cell in ws[1]:
        cell.font      = Font(name="Arial", bold=True, color=HEADER_FONT, size=10)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, col_name in enumerate(GREENHOUSE_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COL_WIDTHS.get(col_name, 20)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    wb.save(str(path))


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Convert a Comeet export to Greenhouse import format."
    )
    parser.add_argument("input",  help="Comeet export file (.numbers / .csv / .xlsx)")
    parser.add_argument("--out",  default=None,
                        help="Output path (default: greenhouse_[name].xlsx next to input)")
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists():
        sys.exit(f"File not found: {input_path}")

    out_path = (
        Path(args.out).expanduser().resolve()
        if args.out
        else input_path.parent / f"greenhouse_{input_path.stem}.xlsx"
    )

    print(f"Reading:      {input_path}")
    df = read_input(input_path)
    print(f"              {len(df)} candidate(s), {len(df.columns)} column(s) found")

    print("Transforming...")
    gh_df = transform(df)

    if "Milestone" in gh_df.columns and "Current stage" in df.columns:
        mapping = {}
        for raw, mapped in zip(df["Current stage"], gh_df["Milestone"]):
            raw_s = str(raw).strip() if raw is not None else ""
            if raw_s and raw_s.lower() not in ("none", "nan"):
                mapping[raw_s] = mapped
        if mapping:
            print("  Milestone mapping:")
            for src, dst in sorted(mapping.items()):
                print(f"    {src!r:35s} → {dst!r}")

    print(f"Writing:      {out_path}")
    gh_df.to_excel(str(out_path), index=False, sheet_name="Sheet people")

    print("Formatting...")
    format_workbook(out_path)

    print(f"\n✓  Done — {len(gh_df)} candidate(s) written to:\n   {out_path}")


if __name__ == "__main__":
    main()
